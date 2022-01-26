import openpyxl
import datetime
import smtplib
import schedule

from mini_Projects.ExcelFile.File2 import xlRead

def sendMail(mailDict):
    server=smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login("Sender Mail ID", "Password")
    
    for key in mailDict:
        message="Wishing You Happy Birthday {} :)".format(mailDict[key])
        server.sendmail("Sender Mail ID", key, message)
        print("\n Mail Send Successfully....")
        server.quit()
    
def readXL():
    path = "file.xlsx"
    
    sendDict = dict()
    nameDict = dict()
    mailDict = dict()
    
    sendMailList = list()
    mailList = list()
    dateLsit = list()
    monthList = list()
    nameList = list()
    
    wb = openpyxl.load_workbook(path)
    sheetObj = wb.active
    
    cols = sheetObj.max_column
    rows = sheetObj.max_row
    
    #dateList 
    for i in range(2,rows+1):
        for j in range(1,2):
            cellObj1 = sheetObj.cell(row=i, column=j)
            dateLsit.append(cellObj1.value)
    
    #monthList
    for i in range(2,rows+1):
        for j in range(2,3):
            cellObj2 = sheetObj.cell(row=i, column=j)
            monthList.append(cellObj2.value)
            
    #mailList
    for i in range(2,rows+1):
        for j in range(4,cols):
            cellObj3 = sheetObj.cell(row=i, column=j)
            mailList.append(cellObj3.value)
            
    #nameList
    for i in range(2,rows+1):
        for j in range(5,cols+1):
            cellObj4 = sheetObj.cell(row=i,column=j)
            nameList.append(cellObj4.value)
    
    #create one dictionary of mailList and dateList as sendDict
    for i in range(len(mailList)):
        sendDict[mailList[i]] = dateLsit[i]
        
    #create one dictionary of nameList and mailList as nameDict
    for i in range(len(nameList)):
        nameDict[mailList[i]] = nameList[i]
        
    
    dateObj = datetime.datetime.now()
    months = dateObj.month
    days = dateObj.day
    
    for i in monthList:
        if months == i:
            for key in sendDict:
                if sendDict[key] == days:
                    sendMailList.append(key)
            break
    
    for key in nameDict:
        for i in range(len(sendMailList)):
            if key == sendMailList[i]:    
                mailDict = { keys: nameDict[keys] for keys in nameDict.keys() & {sendMailList[i]}}
                sendMail(mailDict)
    
def main():
    
    schedule.every().day.at("00:00").do(xlRead)
    
    while True:
        schedule.run_pending()
        
if __name__=="__main__":
    main()